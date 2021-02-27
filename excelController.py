import openpyxl
import datetime

def invoke(itemlist):
    try:
        wb = openpyxl.load_workbook(filename = 'upjong.xlsx', read_only=False, data_only=True)
        sheet = wb.active
        print(sheet.title)
    except:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '2019'
        idx = 1
        for item in itemlist:
            sheet.cell(row = idx, column = 1).value = item
            sheet.cell(row = idx, column = 2).value = itemlist[item]
            idx += 1
        wb.save('upjong.xlsx')