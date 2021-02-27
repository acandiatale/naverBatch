import openpyxl
from openpyxl.utils import get_column_letter
import time

try:
    wb = openpyxl.load_workbook(filename = 'upjong.xlsx', read_only=False, data_only=True)
    sheet = wb.worksheets[1]
    print(sheet.max_row)
    year = 2019
    # if wb.sheetnames[0] == str(year):

    # else:
    #     sheet = wb.create_sheet(str(year+1))

    # if sheet['A1'].value == None :
    #     print("empty!!!")
    # else:
    #     print(sheet['A1'].value)
    #     print(wb['testsheet'])
    # wb.save('upjong.xlsx')\
    # data = []
    # index = 0
    # for row in sheet.rows:
    #     print(row.cell)
except:
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('2019')
    sheet['A1'] = 'hoit'
    wb.save('upjong.xlsx')

